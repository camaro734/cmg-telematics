"""Servicio de datos del reporte de trabajos (intervenciones).

GENÉRICO: las columnas de señal salen de las señales marcadas ``is_report`` en el
``sensor_schema`` del vehicle_type (configurable por cliente desde el editor). El
formato VPS es la primera instancia, no algo hardcodeado.

El reporte SOLO LEE las intervenciones (`work_cycle`); no escribe nada.

Estructura:
  - ``generate_report_data``  — lee de la BD (read-only) y ensambla filas + totales.
  - helpers puros (``_report_signals``, ``_signal_value``, ``_compute_totals``,
    ``compute_leg_km``) — testeables sin BD ni red.
"""
import base64
import logging
import uuid
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Awaitable, Callable

from jinja2 import Environment, FileSystemLoader
from sqlalchemy import text
from sqlalchemy.ext.asyncio import AsyncSession

from app.services.geo import haversine_m
from app.services.geocoding import nominatim_reverse
from app.services.report_branding import build_issuer
from app.services.routing import valhalla_trace_distance_m

logger = logging.getLogger(__name__)

_TEMPLATES_DIR = Path(__file__).parent.parent / "templates"
_LOGO_PATH = Path(__file__).parent.parent / "static" / "logos" / "cmgtrack.png"

# report_aggregate → sufijo de la clave en cycle_data (ver cycle_detector._build_cycle_data).
# max/min/avg vienen de aggregate_fields; "last" es el snapshot de fin (_end).
_AGG_SUFFIX = {"max": "max", "min": "min", "avg": "avg", "last": "end"}

# Tope defensivo de puntos de traza enviados a map-matching por tramo.
_MAX_TRACE_POINTS = 500


def _report_signals(sensor_schema: list | None) -> list[dict]:
    """Señales marcadas is_report en un sensor_schema, con su agregado y etiqueta."""
    out: list[dict] = []
    for s in sensor_schema or []:
        if isinstance(s, dict) and s.get("is_report") and s.get("key"):
            out.append({
                "key": s["key"],
                "label": s.get("label") or s["key"],
                "aggregate": s.get("report_aggregate") or "max",
                "unit": s.get("unit"),
            })
    return out


def _merge_signals(schemas: list[list | None]) -> list[dict]:
    """Une las señales is_report de varios vehicle_types, deduplicando por key."""
    seen: dict[str, dict] = {}
    for schema in schemas:
        for sig in _report_signals(schema):
            seen.setdefault(sig["key"], sig)
    return list(seen.values())


def _signal_value(cycle_data: dict | None, sig: dict) -> float | None:
    """Lee el valor de una señal del reporte desde cycle_data según su agregado."""
    suffix = _AGG_SUFFIX.get(sig["aggregate"], "max")
    val = (cycle_data or {}).get(f'{sig["key"]}_{suffix}')
    try:
        return float(val) if val is not None else None
    except (TypeError, ValueError):
        return None


def _haversine_sum_m(trace: list[tuple[float, float]]) -> float:
    """Suma de distancias haversine entre puntos consecutivos de la traza (metros)."""
    total = 0.0
    for (a_lat, a_lon), (b_lat, b_lon) in zip(trace, trace[1:]):
        total += haversine_m(a_lat, a_lon, b_lat, b_lon)
    return total


async def compute_leg_km(
    trace: list[tuple[float, float]],
    prev_point: tuple[float, float] | None,
    curr_point: tuple[float, float] | None,
    matcher: Callable[[list[tuple[float, float]]], Awaitable[float]] | None = None,
) -> float:
    """Kilómetros del tramo recorrido para llegar a una intervención.

    1) Map-matching de Valhalla sobre la traza GPS del tramo (si hay ≥2 puntos).
    2) Fallback: suma haversine de la traza si el matching falla.
    3) Si no hay traza: haversine en línea recta entre punto anterior y actual.
    """
    matcher = matcher or valhalla_trace_distance_m
    if trace and len(trace) >= 2:
        try:
            return round(await matcher(trace) / 1000.0, 2)
        except Exception as exc:  # noqa: BLE001 — el matching no debe bloquear el reporte
            logger.warning("map-matching falló, usando haversine: %s", exc)
            return round(_haversine_sum_m(trace) / 1000.0, 2)
    if prev_point and curr_point:
        return round(haversine_m(*prev_point, *curr_point) / 1000.0, 2)
    return 0.0


def _compute_totals(rows: list[dict], signals: list[dict]) -> dict:
    """Totales del reporte: nº de intervenciones, km total y min/max por señal."""
    totals: dict[str, Any] = {
        "intervenciones": len(rows),
        "km_total": round(sum(r["kilometraje"] for r in rows), 2),
        "senales": {},
    }
    for sig in signals:
        vals = [r["senales"][sig["key"]] for r in rows if r["senales"].get(sig["key"]) is not None]
        if vals:
            totals["senales"][sig["key"]] = {"min": min(vals), "max": max(vals)}
        else:
            totals["senales"][sig["key"]] = {"min": None, "max": None}
    return totals


def _ot_label(row: dict) -> str:
    """OT mostrada: doc_number solo si la asociación es 'auto'; si no, 'Sin asignar'."""
    if row.get("assignment_status") == "auto" and row.get("doc_number"):
        return row["doc_number"]
    return "Sin asignar"


def _client_label(row: dict) -> str:
    """Cliente: final de la OT, o tenant de la OT; si no hay OT, tenant del vehículo."""
    return (
        row.get("final_client_name")
        or row.get("wo_tenant_name")
        or row.get("vehicle_tenant_name")
        or "—"
    )


async def _fetch_trace(
    db: AsyncSession, vehicle_id, t0: datetime, t1: datetime
) -> list[tuple[float, float]]:
    """Traza GPS (lat, lon) del vehículo en [t0, t1), submuestreada a _MAX_TRACE_POINTS."""
    if t1 <= t0:
        return []
    res = await db.execute(
        text("""
            SELECT lat, lon FROM telemetry_record
            WHERE vehicle_id = :vid AND time >= :t0 AND time < :t1
              AND lat IS NOT NULL AND lon IS NOT NULL
            ORDER BY time
        """),
        {"vid": str(vehicle_id), "t0": t0, "t1": t1},
    )
    pts = [(float(r["lat"]), float(r["lon"])) for r in res.mappings().all()]
    if len(pts) > _MAX_TRACE_POINTS:
        step = len(pts) // _MAX_TRACE_POINTS + 1
        pts = pts[::step]
    return pts


async def _resolve_address(row: dict, cache: dict, reverse_fn) -> str:
    """Dirección destino de la OT si está asociada; si no, geocode inverso del inicio."""
    if row.get("stop_address"):
        return row["stop_address"]
    lat, lon = row.get("lat"), row.get("lon")
    if lat is None or lon is None:
        return "—"
    key = (round(float(lat), 4), round(float(lon), 4))
    if key in cache:
        return cache[key]
    try:
        addr = await reverse_fn(float(lat), float(lon)) or "—"
    except Exception as exc:  # noqa: BLE001 — el geocoder no debe bloquear el reporte
        logger.warning("geocode inverso falló para %s: %s", key, exc)
        addr = "—"
    cache[key] = addr
    return addr


async def generate_report_data(
    db: AsyncSession,
    *,
    from_dt: datetime,
    to_dt: datetime,
    vehicle_id: uuid.UUID | None = None,
    client_id: uuid.UUID | None = None,
    tenant_scope: uuid.UUID | None = None,
    matcher=None,
    reverse_fn=None,
) -> dict:
    """Ensambla el reporte de intervenciones (read-only) para el rango y filtros dados."""
    reverse_fn = reverse_fn or nominatim_reverse

    where = ["wc.started_at >= :from_dt", "wc.started_at < :to_dt"]
    params: dict[str, Any] = {"from_dt": from_dt, "to_dt": to_dt}
    if vehicle_id:
        where.append("wc.vehicle_id = :vid")
        params["vid"] = str(vehicle_id)
    if client_id:
        where.append("v.tenant_id = :cid")
        params["cid"] = str(client_id)
    if tenant_scope is not None:
        where.append("v.tenant_id = :scope")
        params["scope"] = str(tenant_scope)

    sql = f"""
        SELECT wc.id, wc.vehicle_id, wc.started_at, wc.ended_at, wc.duration_seconds,
               wc.lat, wc.lon, wc.cycle_data, wc.assignment_status, wc.work_order_id,
               v.name AS vehicle_name, v.vehicle_type_id,
               vt.sensor_schema,
               vtn.name AS vehicle_tenant_name,
               wo.doc_number, wo.final_client_name,
               wot.name AS wo_tenant_name,
               wos.address AS stop_address
        FROM work_cycle wc
        JOIN vehicle v ON v.id = wc.vehicle_id
        LEFT JOIN vehicle_type vt ON vt.id = v.vehicle_type_id
        LEFT JOIN tenant vtn ON vtn.id = v.tenant_id
        LEFT JOIN work_order wo ON wo.id = wc.work_order_id
        LEFT JOIN tenant wot ON wot.id = wo.tenant_id
        LEFT JOIN work_order_stop wos ON wos.id = wc.work_order_stop_id
        WHERE {" AND ".join(where)}
        ORDER BY wc.vehicle_id, wc.started_at
    """
    raw = [dict(r) for r in (await db.execute(text(sql), params)).mappings().all()]
    return await assemble_report(
        db, raw, from_dt=from_dt, to_dt=to_dt,
        vehicle_id=vehicle_id, client_id=client_id, matcher=matcher, reverse_fn=reverse_fn,
    )


async def assemble_report(
    db: AsyncSession,
    raw: list[dict],
    *,
    from_dt: datetime,
    to_dt: datetime,
    vehicle_id: uuid.UUID | None = None,
    client_id: uuid.UUID | None = None,
    matcher=None,
    reverse_fn=None,
) -> dict:
    """Ensambla filas + totales a partir de las intervenciones crudas (read-only).

    ``raw`` = dicts con las claves del SELECT de ``generate_report_data``. Separado para
    poder ensamblar el reporte a partir de intervenciones en memoria (verificación en frío).
    """
    reverse_fn = reverse_fn or nominatim_reverse

    # Señales del reporte: unión de los is_report de los vehicle_types implicados.
    signals = _merge_signals([r.get("sensor_schema") for r in raw])

    addr_cache: dict = {}
    rows: list[dict] = []
    prev_by_vehicle: dict[str, dict] = {}

    for r in raw:
        vid = str(r["vehicle_id"])
        prev = prev_by_vehicle.get(vid)
        # Traza del tramo: desde el fin de la intervención anterior (o el inicio del
        # rango para la primera) hasta el inicio de la actual.
        leg_start = prev["ended_at"] if (prev and prev.get("ended_at")) else from_dt
        trace = await _fetch_trace(db, r["vehicle_id"], leg_start, r["started_at"])
        prev_point = (
            (float(prev["lat"]), float(prev["lon"]))
            if prev and prev.get("lat") is not None and prev.get("lon") is not None
            else None
        )
        curr_point = (
            (float(r["lat"]), float(r["lon"]))
            if r.get("lat") is not None and r.get("lon") is not None
            else None
        )
        km = await compute_leg_km(trace, prev_point, curr_point, matcher=matcher)

        senales = {sig["key"]: _signal_value(r.get("cycle_data"), sig) for sig in signals}
        rows.append({
            "fecha": r["started_at"].strftime("%d/%m/%Y"),
            "started_at": r["started_at"].isoformat(),
            "ot": _ot_label(r),
            "cliente": _client_label(r),
            "vehiculo": r["vehicle_name"],
            "senales": senales,
            "kilometraje": km,
            "direccion": await _resolve_address(r, addr_cache, reverse_fn),
            "duracion_s": r.get("duration_seconds"),
        })
        prev_by_vehicle[vid] = r

    return {
        "filtros": {
            "desde": from_dt.isoformat(),
            "hasta": to_dt.isoformat(),
            "vehicle_id": str(vehicle_id) if vehicle_id else None,
            "client_id": str(client_id) if client_id else None,
        },
        "columnas_senal": [
            {"key": s["key"], "label": s["label"], "unit": s.get("unit")} for s in signals
        ],
        "filas": rows,
        "totales": _compute_totals(rows, signals),
    }


def _logo_b64() -> str | None:
    """Logo CMG en base64 para incrustar en el PDF (None si falta el fichero)."""
    try:
        return base64.b64encode(_LOGO_PATH.read_bytes()).decode()
    except Exception:  # noqa: BLE001 — la ausencia de logo no debe romper el PDF
        return None


def _periodo_label(report: dict) -> str:
    """'dd/mm/aaaa – dd/mm/aaaa' a partir de los filtros del reporte."""
    desde = datetime.fromisoformat(report["filtros"]["desde"]).strftime("%d/%m/%Y")
    hasta = datetime.fromisoformat(report["filtros"]["hasta"]).strftime("%d/%m/%Y")
    return f"{desde} – {hasta}"


def render_report_pdf(
    report: dict,
    *,
    title: str = "Parte de trabajos",
    subtitle: str = "",
    issuer: dict | None = None,
) -> bytes:
    """Renderiza el reporte a PDF (formato VPS) con WeasyPrint. Función síncrona.

    ``issuer`` = membrete del emisor (build_issuer(tenant)); si None, cae a CMG.
    """
    import weasyprint

    if issuer is None:
        issuer = build_issuer(None)
    env = Environment(loader=FileSystemLoader(str(_TEMPLATES_DIR)), autoescape=True)
    template = env.get_template("reports/work_cycle_report.html")
    html = template.render(
        report=report,
        title=title,
        subtitle=subtitle,
        issuer=issuer,
        periodo=_periodo_label(report),
        generated_at=datetime.now(timezone.utc).strftime("%d/%m/%Y %H:%M"),
    )
    return weasyprint.HTML(string=html, base_url=str(_TEMPLATES_DIR)).write_pdf()


def _col_header(col: dict) -> str:
    return f"{col['label']} ({col['unit']})" if col.get("unit") else col["label"]


def render_report_xlsx(report: dict) -> bytes:
    """Renderiza el reporte a Excel (.xlsx) con openpyxl. Función síncrona."""
    from io import BytesIO

    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    cols = report["columnas_senal"]
    wb = Workbook()
    ws = wb.active
    ws.title = "Intervenciones"

    headers = ["Fecha", "OT", "Cliente"] + [_col_header(c) for c in cols] + ["Kilometraje (km)", "Dirección"]
    ws.append(headers)
    header_fill = PatternFill("solid", fgColor="1A2E4A")
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for fila in report["filas"]:
        row = [fila["fecha"], fila["ot"], fila["cliente"]]
        for c in cols:
            v = fila["senales"].get(c["key"])
            row.append(v if v is not None else "—")
        row += [fila["kilometraje"], fila["direccion"]]
        ws.append(row)

    # Fila de totales
    totals = report["totales"]
    total_row = [f"TOTALES ({totals['intervenciones']} interv.)", "", ""]
    for c in cols:
        t = totals["senales"].get(c["key"], {})
        total_row.append(
            f"{t['min']:.2f} / {t['max']:.2f}" if t.get("min") is not None else "—"
        )
    total_row += [totals["km_total"], ""]
    ws.append(total_row)
    for cell in ws[ws.max_row]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="EEF2F7")

    # Anchos de columna razonables
    widths = [12, 16, 22] + [16] * len(cols) + [16, 60]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
