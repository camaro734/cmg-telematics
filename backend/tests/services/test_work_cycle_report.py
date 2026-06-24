"""Tests de los helpers puros del servicio de datos del reporte de trabajos.

No tocan BD ni red: validan la extracción de señales is_report, el agregado leído
de cycle_data, los km del tramo (map-matching + fallback haversine) y los totales.
"""
import pytest

from app.services.work_cycle_report import (
    _AGG_SUFFIX,
    _client_label,
    _compute_totals,
    _merge_signals,
    _ot_label,
    _report_signals,
    _signal_value,
    compute_leg_km,
)

SCHEMA = [
    {"key": "min_depresor", "label": "Presión depresor", "unit": "bar",
     "is_report": True, "report_aggregate": "max"},
    {"key": "presion_agua", "label": "Presión agua", "is_report": True, "report_aggregate": "avg"},
    {"key": "rpm", "label": "RPM", "is_report": False},   # no marcada
    {"is_report": True},                                   # sin key → ignorada
]


def test_report_signals_filters_is_report():
    sigs = _report_signals(SCHEMA)
    keys = [s["key"] for s in sigs]
    assert keys == ["min_depresor", "presion_agua"]
    assert sigs[0]["aggregate"] == "max" and sigs[0]["label"] == "Presión depresor"
    assert sigs[1]["aggregate"] == "avg"


def test_report_signals_default_aggregate_is_max():
    sigs = _report_signals([{"key": "x", "is_report": True}])
    assert sigs[0]["aggregate"] == "max"


def test_merge_signals_dedup_by_key():
    other = [{"key": "min_depresor", "is_report": True, "report_aggregate": "min"}]
    merged = _merge_signals([SCHEMA, other])
    keys = [s["key"] for s in merged]
    assert keys == ["min_depresor", "presion_agua"]   # min_depresor no se duplica


def test_signal_value_maps_aggregate_to_cycle_data_key():
    cd = {"min_depresor_max": 24.0, "presion_agua_avg": 3.5, "min_depresor_min": 3.0}
    assert _signal_value(cd, {"key": "min_depresor", "aggregate": "max"}) == 24.0
    assert _signal_value(cd, {"key": "presion_agua", "aggregate": "avg"}) == 3.5
    assert _signal_value(cd, {"key": "min_depresor", "aggregate": "min"}) == 3.0


def test_signal_value_last_maps_to_end_snapshot():
    assert _AGG_SUFFIX["last"] == "end"
    cd = {"nivel_end": 12.0}
    assert _signal_value(cd, {"key": "nivel", "aggregate": "last"}) == 12.0


def test_signal_value_missing_key_is_none():
    assert _signal_value({}, {"key": "x", "aggregate": "max"}) is None
    assert _signal_value(None, {"key": "x", "aggregate": "max"}) is None


def test_ot_label_only_when_auto():
    assert _ot_label({"assignment_status": "auto", "doc_number": "PT-2026-00001"}) == "PT-2026-00001"
    assert _ot_label({"assignment_status": "pending", "doc_number": "PT-2026-00001"}) == "Sin asignar"
    assert _ot_label({"assignment_status": "sin_asignar", "doc_number": None}) == "Sin asignar"


def test_client_label_priority():
    assert _client_label({"final_client_name": "ACME", "wo_tenant_name": "T", "vehicle_tenant_name": "V"}) == "ACME"
    assert _client_label({"wo_tenant_name": "T", "vehicle_tenant_name": "V"}) == "T"
    assert _client_label({"vehicle_tenant_name": "V"}) == "V"
    assert _client_label({}) == "—"


@pytest.mark.asyncio
async def test_compute_leg_km_uses_matcher():
    async def fake_matcher(trace):
        return 5000.0   # 5 km en metros
    km = await compute_leg_km([(39.5, -0.4), (39.51, -0.4)], None, None, matcher=fake_matcher)
    assert km == 5.0


@pytest.mark.asyncio
async def test_compute_leg_km_falls_back_to_haversine_on_matcher_error():
    async def boom(trace):
        raise RuntimeError("valhalla caído")
    # ~1.11 km entre los dos puntos (0.01° de latitud)
    km = await compute_leg_km([(39.50, -0.40), (39.51, -0.40)], None, None, matcher=boom)
    assert 1.0 < km < 1.3


@pytest.mark.asyncio
async def test_compute_leg_km_straight_line_when_no_trace():
    km = await compute_leg_km([], (39.50, -0.40), (39.51, -0.40), matcher=None)
    assert 1.0 < km < 1.3


@pytest.mark.asyncio
async def test_compute_leg_km_zero_when_nothing():
    assert await compute_leg_km([], None, None, matcher=None) == 0.0


def _sample_report(filas):
    from app.services.work_cycle_report import _compute_totals
    signals = [{"key": "min_depresor", "label": "Min Depresor", "unit": "Min", "aggregate": "max"}]
    return {
        "filtros": {"desde": "2026-06-22T00:00:00+00:00", "hasta": "2026-06-23T00:00:00+00:00",
                    "vehicle_id": None, "client_id": None},
        "columnas_senal": [{"key": "min_depresor", "label": "Min Depresor", "unit": "Min"}],
        "filas": filas,
        "totales": _compute_totals(filas, signals),
    }


def test_render_report_pdf_produces_pdf_bytes():
    from app.services.work_cycle_report import render_report_pdf
    report = _sample_report([
        {"fecha": "22/06/2026", "ot": "Sin asignar", "cliente": "DELIMEX",
         "senales": {"min_depresor": 24.0}, "kilometraje": 12.3, "direccion": "Calle Test, Valencia"},
    ])
    pdf = render_report_pdf(report, subtitle="test")
    assert pdf[:4] == b"%PDF"
    assert len(pdf) > 1000


def test_render_report_pdf_empty_is_valid():
    from app.services.work_cycle_report import render_report_pdf
    pdf = render_report_pdf(_sample_report([]))
    assert pdf[:4] == b"%PDF"


def test_render_report_xlsx_has_headers_rows_and_totals():
    from io import BytesIO
    from openpyxl import load_workbook
    from app.services.work_cycle_report import render_report_xlsx

    report = _sample_report([
        {"fecha": "22/06/2026", "ot": "PT-2026-00001", "cliente": "DELIMEX",
         "senales": {"min_depresor": 24.0}, "kilometraje": 12.3, "direccion": "Calle Test"},
        {"fecha": "22/06/2026", "ot": "Sin asignar", "cliente": "DELIMEX",
         "senales": {"min_depresor": None}, "kilometraje": 4.0, "direccion": "Otra"},
    ])
    xlsx = render_report_xlsx(report)
    ws = load_workbook(BytesIO(xlsx)).active
    rows = list(ws.iter_rows(values_only=True))
    assert rows[0] == ("Fecha", "OT", "Cliente", "Min Depresor (Min)", "Kilometraje (km)", "Dirección")
    assert rows[1][1] == "PT-2026-00001" and rows[1][3] == 24.0
    assert rows[2][3] == "—"   # señal ausente
    assert rows[-1][0].startswith("TOTALES")
    assert rows[-1][4] == 16.3   # km total


def test_compute_totals():
    signals = [{"key": "min_depresor", "aggregate": "max"}]
    rows = [
        {"kilometraje": 10.0, "senales": {"min_depresor": 24.0}},
        {"kilometraje": 5.5, "senales": {"min_depresor": 16.0}},
        {"kilometraje": 2.0, "senales": {"min_depresor": None}},
    ]
    totals = _compute_totals(rows, signals)
    assert totals["intervenciones"] == 3
    assert totals["km_total"] == 17.5
    assert totals["senales"]["min_depresor"] == {"min": 16.0, "max": 24.0}
